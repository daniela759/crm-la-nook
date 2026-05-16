"""
Extrage datele Nook din XLSX-urile primite și produce un JSON normalizat.
- Acorduri semnate → contacte + copii (cu telefon, email, vârstă)
- Vizite mai/aprilie/martie → lead-uri (VISIT, status PRESENT) + tranzacții
- Pachete 4/8 intrări → abonamente

Numele clienților din vizite se face match cu contactele din Acorduri
folosind normalizare: lowercase, fără diacritice, cuvinte sortate alfabetic.
Vizitele fără match primesc un contact nou (sursă = Walk-in).
"""

import json
import re
import sys
import unicodedata
from datetime import datetime, date
from pathlib import Path
import openpyxl

ROOT = Path("/Users/danielasuhanea/Downloads")
ACORDURI = ROOT / "Playtime sessions .xlsx"
OUT = Path("/Users/danielasuhanea/Projects/nook-crm/prisma/import/data.json")
REF_TODAY = date(2026, 5, 16)  # consistent cu memory `currentDate`

# ─── Normalizare nume ─────────────────────────────────────────────────────
def strip_diacritics(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))

def normalize_name(name) -> str:
    if not name: return ""
    s = strip_diacritics(str(name)).lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    parts = sorted(s.split())
    return " ".join(parts)

def split_name(name: str):
    """Întoarce (firstName, lastName). Convenția în date: «Nume Prenume»."""
    parts = re.split(r"\s+", str(name).strip())
    parts = [p for p in parts if p]
    if not parts: return ("", "")
    if len(parts) == 1: return (parts[0], "")
    # Heuristic: în "Acorduri semnate" e "Nume Prenume", dar în vizite uneori "Prenume Nume".
    # Folosim ordinea originală: primul = first, restul = last
    return (parts[0], " ".join(parts[1:]))

# ─── Parsare vârstă text → data nașterii aproximativă ─────────────────────
AGE_PATTERN = re.compile(
    r"(?:(\d+)\s*ani?)?\s*(?:și\s*)?(?:(\d+)\s*lun[ai])?",
    re.IGNORECASE,
)

def parse_age_to_birth(age_text):
    if not age_text: return None
    s = str(age_text).strip()
    if not s: return None
    m = AGE_PATTERN.search(s)
    if not m: return None
    years = int(m.group(1) or 0)
    months = int(m.group(2) or 0)
    if years == 0 and months == 0: return None
    # Calculează data nașterii aproximativ
    total_months = years * 12 + months
    birth_year = REF_TODAY.year
    birth_month = REF_TODAY.month - total_months
    while birth_month <= 0:
        birth_month += 12
        birth_year -= 1
    birth_day = max(1, min(REF_TODAY.day, 28))
    return f"{birth_year:04d}-{birth_month:02d}-{birth_day:02d}"

# ─── Parsare telefon (vine ca number din Excel) ───────────────────────────
def parse_phone(raw):
    if raw is None or raw == "-": return None
    if isinstance(raw, (int, float)):
        s = str(int(raw))
        # Adăugăm 0 inițial dacă lipsește
        if not s.startswith("0"): s = "0" + s
        return s
    s = str(raw).strip()
    if not s or s == "-": return None
    return s

# ─── Parsare data ─────────────────────────────────────────────────────────
def parse_date_dotted(raw):
    if raw is None: return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = str(raw).strip()
    # "30.03.2026"
    m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None

# ─── Parsare interval sesiune (din coloana A din sheet-urile vizite) ──────
def parse_interval(raw) -> str:
    """Returnează HH:MM ca string."""
    if raw is None: return "10:00"
    # Excel transformă "10-12" în datetime(2026, 10, 12)
    if isinstance(raw, datetime):
        # E "10-12" → sesiunea de dimineață
        return "10:00"
    s = str(raw).strip()
    if "17" in s and "19" in s:
        return "17:00"
    if "10" in s and "12" in s:
        return "10:00"
    # Fallback
    return "10:00"

# ─── Mapează metoda de plată ──────────────────────────────────────────────
def classify_payment(raw, price):
    """
    Returnează (revenue_type, payment_method, from_subscription, sub_type_hint)
    - from_subscription = True dacă vizita consumă din abonament
    """
    s = (str(raw).strip().lower() if raw else "")
    if "pachet" in s:
        sub_type = "ENTRIES_8" if "8" in s else "ENTRIES_4"
        # Dacă preț = 0, e consumarea unei intrări
        if price is None or price == 0:
            return ("VISIT", "CARD", True, sub_type)
        # Dacă preț > 0, e cumpărarea abonamentului
        return ("SUBSCRIPTION", "CARD", False, sub_type)
    if "card" in s:
        return ("VISIT", "CARD", False, None)
    if "cash" in s:
        return ("VISIT", "CASH", False, None)
    if "platit" in s or "la locatie" in s or "locație" in s or "locatie" in s:
        return ("VISIT", "CASH", False, None)
    return ("VISIT", "CASH", False, None)

# ─── Extragere contacte ───────────────────────────────────────────────────
def extract_contacts(wb):
    ws = wb["Acorduri semnate"]
    contacts = {}  # key normalizat → contact dict
    duplicates = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name_adult, phone_raw, email_raw, name_child, age_text, sign_date_raw, photo_raw, active_raw = row[:8]
        if not name_adult: continue
        # Sărim rândurile fără telefon ȘI fără email — nu putem importa fără date
        phone = parse_phone(phone_raw)
        email = (str(email_raw).strip() if email_raw and email_raw != "-" else None)

        first, last = split_name(name_adult)
        norm = normalize_name(name_adult)
        photo_agreement = (str(photo_raw).strip().lower() == "da") if photo_raw else False
        active = (str(active_raw).strip().lower() != "inactiv") if active_raw else True
        sign_date = parse_date_dotted(sign_date_raw)

        if norm in contacts:
            # Adultul a fost deja întâlnit — adăugăm copilul la el
            if name_child:
                contacts[norm]["children"].append({
                    "name": str(name_child).strip(),
                    "birthDate": parse_age_to_birth(age_text),
                    "age_text": age_text,
                })
            duplicates += 1
            continue

        contacts[norm] = {
            "norm": norm,
            "firstName": first,
            "lastName": last,
            "phone": phone,
            "email": email,
            "signDate": sign_date,
            "photoAgreement": photo_agreement,
            "active": active,
            "children": [],
        }
        if name_child:
            contacts[norm]["children"].append({
                "name": str(name_child).strip(),
                "birthDate": parse_age_to_birth(age_text),
                "age_text": age_text,
            })

    print(f"  Contacte unice: {len(contacts)} (din care {duplicates} rânduri secundare)", file=sys.stderr)
    return contacts

# ─── Extragere vizite dintr-un sheet ──────────────────────────────────────
def extract_visits(ws, month_label: str):
    visits = []
    current_date = None
    current_interval = "10:00"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        col_interval, col_name, col_kids, col_adults, col_price, col_payment = row[:6]
        col_notes = row[6] if len(row) > 6 else None

        # Header de zi: prima coloană e dată ca text "DD.MM.YYYY"
        if col_interval and not col_name and not col_kids:
            d = parse_date_dotted(col_interval)
            if d:
                current_date = d
                current_interval = "10:00"  # default începutul zilei
            continue

        # Determinăm intervalul de timp (carry-over din celulele anterioare)
        if col_interval is not None:
            current_interval = parse_interval(col_interval)

        if not col_name: continue
        if str(col_name).strip().lower() == "nimeni": continue

        kids = int(col_kids) if col_kids else 0
        adults = int(col_adults) if col_adults else 0
        price = float(col_price) if col_price is not None else None

        rev, pm, from_sub, sub_hint = classify_payment(col_payment, price)
        # Dacă rev=SUBSCRIPTION, e vânzare abonament — îl marcăm separat
        is_subscription_sale = (rev == "SUBSCRIPTION")

        visits.append({
            "date": current_date,
            "time": current_interval,
            "contact_norm": normalize_name(col_name),
            "contact_raw_name": str(col_name).strip(),
            "kids": kids,
            "adults": adults,
            "price": price if price is not None else 0,
            "payment_raw": col_payment,
            "from_subscription": from_sub,
            "is_subscription_sale": is_subscription_sale,
            "sub_type_hint": sub_hint,
            "notes": str(col_notes).strip() if col_notes else None,
            "month": month_label,
        })

    return visits

# ─── Extragere pachete abonamente ─────────────────────────────────────────
def extract_packages(wb):
    subs = []

    # Pachete 4 intrari
    ws = wb["Pachete 4 intrari"]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        # col_a = lună (header) sau None
        _, holder, purchase_raw, expire_raw, e1, e2, e3, e4, active = row[:9]
        if not holder: continue
        purchased = parse_date_dotted(purchase_raw)
        if not purchased: continue
        used = sum(1 for e in [e1, e2, e3, e4] if e is True)
        subs.append({
            "contact_norm": normalize_name(holder),
            "contact_raw": str(holder).strip(),
            "type": "ENTRIES_4",
            "purchasedAt": purchased,
            "totalEntries": 4,
            "usedEntries": used,
            "pricePaid": 200,  # din date e 200 lei, nu 250 (vechi/discount)
            "expiresAt": parse_date_dotted(expire_raw),
        })

    # Pachete 8 intrari
    ws = wb["Pachete 8 intrari"]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        holder, purchase_raw, expire_raw, e1, e2, e3, e4, e5, e6, e7, e8, active = row[:12]
        if not holder: continue
        purchased = parse_date_dotted(purchase_raw)
        if not purchased: continue
        used = sum(1 for e in [e1, e2, e3, e4, e5, e6, e7, e8] if e is True)
        subs.append({
            "contact_norm": normalize_name(holder),
            "contact_raw": str(holder).strip(),
            "type": "ENTRIES_8",
            "purchasedAt": purchased,
            "totalEntries": 8,
            "usedEntries": used,
            "pricePaid": 350,
            "expiresAt": parse_date_dotted(expire_raw),
        })

    print(f"  Pachete: {len(subs)} (4-intrări + 8-intrări)", file=sys.stderr)
    return subs

# ─── Main ────────────────────────────────────────────────────────────────
def main():
    print("Citesc Playtime sessions .xlsx...", file=sys.stderr)
    wb = openpyxl.load_workbook(ACORDURI, data_only=True)

    print("Extrag contacte din 'Acorduri semnate'...", file=sys.stderr)
    contacts = extract_contacts(wb)

    print("Extrag vizite (martie + aprilie + mai)...", file=sys.stderr)
    all_visits = []
    for sheet, label in [("Vizite martie", "martie"), ("Vizite aprilie", "aprilie"), ("Vizite mai", "mai")]:
        if sheet not in wb.sheetnames: continue
        vs = extract_visits(wb[sheet], label)
        all_visits.extend(vs)
        print(f"  {label}: {len(vs)} vizite", file=sys.stderr)
    print(f"  Total vizite: {len(all_visits)}", file=sys.stderr)

    print("Extrag pachete abonamente...", file=sys.stderr)
    subs = extract_packages(wb)

    # Pentru vizitele cu un nume necunoscut, agregăm match-urile lipsă
    unmatched = {}
    for v in all_visits:
        if v["contact_norm"] not in contacts:
            unmatched[v["contact_norm"]] = unmatched.get(v["contact_norm"], 0) + 1
    print(f"  Vizite cu contact necunoscut (vor primi contact nou): {sum(unmatched.values())} pe {len(unmatched)} persoane", file=sys.stderr)

    # Adăugăm și contactele care apar doar în pachete/vizite, fără acord
    for v in all_visits:
        if v["contact_norm"] not in contacts:
            first, last = split_name(v["contact_raw_name"])
            contacts[v["contact_norm"]] = {
                "norm": v["contact_norm"],
                "firstName": first,
                "lastName": last,
                "phone": None,
                "email": None,
                "signDate": None,
                "photoAgreement": False,
                "active": True,
                "children": [],
                "source_note": "Import — fără acord semnat",
            }
    for s in subs:
        if s["contact_norm"] not in contacts:
            first, last = split_name(s["contact_raw"])
            contacts[s["contact_norm"]] = {
                "norm": s["contact_norm"],
                "firstName": first,
                "lastName": last,
                "phone": None,
                "email": None,
                "signDate": None,
                "photoAgreement": False,
                "active": True,
                "children": [],
                "source_note": "Import — abonament fără acord semnat",
            }

    print(f"  Total contacte finale: {len(contacts)}", file=sys.stderr)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "contacts": list(contacts.values()),
        "visits": all_visits,
        "subscriptions": subs,
    }, ensure_ascii=False, indent=2, default=str))
    print(f"Scris: {OUT}", file=sys.stderr)

if __name__ == "__main__":
    main()
